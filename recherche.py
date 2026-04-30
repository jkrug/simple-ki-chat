#!/usr/bin/env python3
"""Interaktiver Mail-Recherche-Agent für juristische Aktenrecherche.

Orchestriert qmd-Suchen, bewertet Treffer per LLM-Hypothese, baut
inkrementell eine Akte auf. Sessions werden persistent gespeichert.
"""
import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path

import qmd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import readline
    HAVE_READLINE = True
except ImportError:
    HAVE_READLINE = False

COMMANDS = [
    "/search", "/add", "/suggest", "/review", "/review fast", "/reject",
    "/list", "/sessions", "/gaps", "/devil", "/context", "/context edit",
    "/validate-context", "/validate", "/summary", "/export", "/dossier",
    "/akte", "/case", "/edit", "/undo", "/help", "/quit",
]


def setup_readline(history_path: Path) -> None:
    if not HAVE_READLINE:
        return
    history_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        readline.read_history_file(str(history_path))
    except (OSError, FileNotFoundError):
        pass
    readline.set_history_length(1000)

    # Whole line als Match-Einheit — sonst zerlegt readline bei Spaces
    # und Multi-Word-Commands ('/review fast') werden umständlich.
    readline.set_completer_delims("")

    def completer(text: str, state: int):
        if not text.startswith("/"):
            return None
        matches = [c for c in COMMANDS if c.startswith(text)]
        return matches[state] if state < len(matches) else None

    readline.set_completer(completer)
    # macOS Python ist meist gegen libedit gelinkt, nicht GNU readline
    if "libedit" in (readline.__doc__ or ""):
        readline.parse_and_bind("bind ^I rl_complete")
    else:
        readline.parse_and_bind("tab: complete")

    import atexit
    atexit.register(lambda: _safe_write_history(history_path))


def _safe_write_history(path: Path) -> None:
    try:
        readline.write_history_file(str(path))
    except OSError:
        pass

OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://localhost:11434")
DEFAULT_MODEL = os.environ.get("OLLAMA_MODEL", "qwen2.5:32b")
DEFAULT_OUT = Path.home() / "marmalade-fall" / "output"
SESSIONS_DIR = Path(
    os.environ.get("CHATBOT_SESSIONS_DIR")
    or (Path(os.environ.get("XDG_DATA_HOME", Path.home() / ".local" / "share"))
        / "mini-chatbot" / "sessions")
)

SYSTEM_PROMPT = """Du bist ein juristischer Recherche-Assistent. Du hilfst
Joscha, aus seinen privaten E-Mails einen zeitlichen Ablauf zu rekonstruieren.

Wichtig:
- Joscha kennt den Ablauf NICHT mehr. Stelle keine Bestätigungsfragen
  ("Stimmt es, dass am X passiert ist?") - er weiß es nicht. Bilde
  Hypothesen aus den Mailinhalten und stelle sie zur Diskussion.
- Datum, Absender, Empfänger immer aus dem YAML-Frontmatter, nie aus
  dem Mail-Text raten.
- Wenn ein Feld unklar ist: leerlassen oder "?". Lieber nachfragen als
  raten.
- Antworte präzise auf Deutsch, sachlich, juristisch nüchtern.
- Wenn JSON gefordert ist: NUR valides JSON, keine Einleitung.
- Hintergrund-Notizen vom Mandanten sind dessen SUBJEKTIVE ERINNERUNG
  (z. T. Jahre alt) und können fehlerhaft sein. Wenn eine Mail dem
  widerspricht, folge der Mail und weise auf den Widerspruch hin.
  Mail-Belege schlagen Erinnerung."""


# ───────────── Ollama ─────────────

def _post_chat(model: str, messages: list[dict], *, stream: bool,
               json_mode: bool, num_ctx: int):
    payload = {
        "model": model,
        "messages": messages,
        "stream": stream,
        "options": {"num_ctx": num_ctx},
    }
    if json_mode:
        payload["format"] = "json"
    req = urllib.request.Request(
        f"{OLLAMA_URL}/api/chat",
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    return urllib.request.urlopen(req)


def chat(model: str, messages: list[dict], *, json_mode: bool = False,
         num_ctx: int = 32768) -> str:
    with _post_chat(model, messages, stream=False, json_mode=json_mode,
                    num_ctx=num_ctx) as resp:
        obj = json.loads(resp.read())
    return obj.get("message", {}).get("content", "")


def chat_stream(model: str, messages: list[dict], *, num_ctx: int = 32768) -> str:
    out = []
    with _post_chat(model, messages, stream=True, json_mode=False,
                    num_ctx=num_ctx) as resp:
        for line in resp:
            if not line.strip():
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            chunk = obj.get("message", {}).get("content", "")
            if chunk:
                sys.stdout.write(chunk)
                sys.stdout.flush()
                out.append(chunk)
            if obj.get("done"):
                break
    print()
    return "".join(out)


# ───────────── Session ─────────────

def session_path(name: str) -> Path:
    safe = re.sub(r"[^\w\-.]", "_", name)
    return SESSIONS_DIR / f"recherche_{safe}.json"


def load_session(name: str) -> dict:
    p = session_path(name)
    if p.exists():
        return json.loads(p.read_text())
    return {
        "name": name,
        "case_description": "",
        "executed_searches": [],
        "candidates": {},
        "last_suggestions": [],
    }


def list_sessions() -> list[dict]:
    """Liefert Liste aller recherche_*-Sessions im SESSIONS_DIR."""
    if not SESSIONS_DIR.exists():
        return []
    out = []
    for p in SESSIONS_DIR.glob("recherche_*.json"):
        try:
            data = json.loads(p.read_text())
        except (OSError, json.JSONDecodeError):
            continue
        counts = {"accepted": 0, "pending": 0, "rejected": 0}
        for c in data.get("candidates", {}).values():
            counts[c.get("status", "pending")] = counts.get(
                c.get("status", "pending"), 0) + 1
        out.append({
            "name": data.get("name", p.stem.removeprefix("recherche_")),
            "path": p,
            "mtime": p.stat().st_mtime,
            "case": data.get("case_description", ""),
            "counts": counts,
        })
    return sorted(out, key=lambda s: s["mtime"], reverse=True)


def print_sessions() -> None:
    sessions = list_sessions()
    if not sessions:
        print(f"Keine Sessions in {SESSIONS_DIR}.")
        return
    print(f"\n{len(sessions)} Session(s) in {SESSIONS_DIR}:\n")
    for s in sessions:
        ts = datetime.fromtimestamp(s["mtime"]).strftime("%Y-%m-%d %H:%M")
        c = s["counts"]
        print(f"  {s['name']}")
        print(f"    zuletzt: {ts}  |  ✓{c['accepted']} ?{c['pending']} ✗{c['rejected']}")
        if s["case"]:
            print(f"    Fall:    {s['case'][:80]}")
        print()


def save_session(state: dict) -> None:
    SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
    session_path(state["name"]).write_text(
        json.dumps(state, ensure_ascii=False, indent=2)
    )


# ───────────── Kontext-Datei ─────────────

def context_path(state: dict, out_dir: str) -> Path:
    """Pfad zur Kontext-Markdown-Datei. Default: <out_dir>/kontext.md."""
    p = state.get("context_file") or ""
    if p:
        return Path(p).expanduser()
    return Path(out_dir).expanduser() / "kontext.md"


def load_context_text(state: dict, out_dir: str) -> str:
    p = context_path(state, out_dir)
    try:
        return p.read_text()
    except (OSError, FileNotFoundError):
        return ""


def context_block(state: dict, out_dir: str) -> str:
    """Block für User-Prompts. Leer wenn keine Kontext-Datei vorhanden."""
    text = load_context_text(state, out_dir).strip()
    if not text:
        return ""
    return ("\n### Hintergrund vom Mandanten (kontextuelle Notizen, "
            "Personen, subjektive Sicht — KEIN Mailinhalt):\n"
            f"{text}\n")


def cmd_context(state: dict, arg: str, out_dir: str) -> None:
    arg = arg.strip()
    p = context_path(state, out_dir)
    if not arg:
        if p.exists():
            txt = p.read_text()
            lines = txt.count("\n") + 1
            print(f"Datei: {p}  ({len(txt)} Bytes, {lines} Zeilen)")
            print("─" * 70)
            preview = txt[:1500]
            print(preview + ("\n…" if len(txt) > 1500 else ""))
        else:
            print(f"Noch keine Kontext-Datei. Default-Pfad: {p}\n"
                  f"  /context edit          öffnet $EDITOR (oder nano/vim)\n"
                  f"  /context <pfad.md>     wechselt auf eine andere Datei")
        return
    if arg == "edit":
        editor = os.environ.get("EDITOR", "")
        if not editor:
            for cand in ("nano", "vim", "vi"):
                if shutil.which(cand):
                    editor = cand
                    break
        if not editor:
            print("Kein $EDITOR gesetzt und nano/vim/vi nicht gefunden.")
            return
        p.parent.mkdir(parents=True, exist_ok=True)
        if not p.exists():
            p.write_text(
                "# Fallkontext\n\n"
                "## Wer ist wer\n\n"
                "## Hintergrund\n\n"
                "## Mein Ablauf in Gedanken\n\n"
                "## Worauf soll geachtet werden\n\n"
            )
        subprocess.run([editor, str(p)])
        return
    new = Path(arg).expanduser()
    state["context_file"] = str(new)
    print(f"Kontext-Pfad gesetzt auf: {new}")
    if not new.exists():
        print("(Datei existiert noch nicht — anlegen mit /context edit.)")


# ───────────── Candidates ─────────────

def make_candidate(uri: str) -> dict:
    text = qmd.fetch(uri, full=True)
    meta, body = qmd.parse_frontmatter(text)
    parts = meta.get("participants", [])
    if isinstance(parts, str):
        parts = [parts]
    return {
        "uri": uri,
        "status": "pending",
        "confidential": False,
        "subject": meta.get("subject", ""),
        "participants": parts,
        "date_start": qmd.normalize_date(meta.get("date_start", "")),
        "date_last": qmd.normalize_date(meta.get("date_last", "")),
        "message_count": meta.get("message_count", "1"),
        "body_excerpt": body[:2500],
        "summary": "",
        "rejection_reason": "",
    }


def status_counts(state: dict) -> dict:
    counts = {"accepted": 0, "pending": 0, "rejected": 0}
    for c in state["candidates"].values():
        counts[c["status"]] = counts.get(c["status"], 0) + 1
    return counts


# ───────────── Commands ─────────────

def cmd_search(state: dict, query: str, top_k: int) -> None:
    if not query:
        print("Suchbegriff fehlt.")
        return
    print(f"→ qmd-Suche: {query!r}")
    try:
        uris = qmd.search(query, top_k, all_results=True, min_score=0.2)
    except subprocess.CalledProcessError as e:
        print(f"qmd-Fehler: {e.stderr}")
        return
    new = 0
    for u in uris:
        if u not in state["candidates"]:
            state["candidates"][u] = make_candidate(u)
            new += 1
    state["executed_searches"].append(query)
    print(f"   {len(uris)} Treffer ({new} neu, {len(uris) - new} bekannt).")


def cmd_add(state: dict, arg: str) -> None:
    """URIs manuell hinzufügen. Format:
        /add qmd://threads/a.md qmd://threads/b.md
        /add @/pfad/zu/datei.txt    (jede Zeile darf eine URI enthalten)
    """
    if not arg:
        print("Usage: /add <uri> [<uri> ...]   oder   /add @datei")
        return
    raw_lines: list[str] = []
    for tok in arg.split():
        if tok.startswith("@"):
            path = Path(tok[1:]).expanduser()
            if not path.exists():
                print(f"Datei nicht gefunden: {path}")
                continue
            raw_lines.extend(path.read_text().splitlines())
        else:
            raw_lines.append(tok)
    uris, seen = [], set()
    for line in raw_lines:
        m = qmd.URI_RE.search(line)
        if m and m.group(0) not in seen:
            seen.add(m.group(0))
            uris.append(m.group(0))
    if not uris:
        print("Keine qmd://…-URIs erkannt.")
        return
    new = 0
    for u in uris:
        if u not in state["candidates"]:
            try:
                state["candidates"][u] = make_candidate(u)
                new += 1
            except subprocess.CalledProcessError as e:
                print(f"qmd get fehlgeschlagen für {u}: {e.stderr}")
    print(f"   {len(uris)} URIs verarbeitet ({new} neu, "
          f"{len(uris) - new} bereits bekannt).")


def cmd_suggest(state: dict, model: str, out_dir: str) -> None:
    accepted = [c for c in state["candidates"].values() if c["status"] == "accepted"]
    timeline = "\n".join(
        f"- {c['date_start']}: {c['subject']} — {c.get('summary', '')}"
        for c in sorted(accepted, key=lambda x: x["date_start"])
    ) or "(noch keine bestätigten Mails)"
    searches = "\n".join(f"- {s}" for s in state["executed_searches"]) or "(noch keine)"
    prompt = f"""Fall: {state['case_description']}
{context_block(state, out_dir)}
Bestätigte Mails (chronologisch):
{timeline}

Bereits ausgeführte Suchen:
{searches}

Schlage 3-5 weitere qmd-Suchanfragen vor, um den Ablauf zu vervollständigen.

WICHTIG zum Format der Suchanfragen:
- NUR reiner Stichworttext (2-5 Wörter, deutsche Begriffe oder Eigennamen).
- KEINE Boolean-Operatoren (AND/OR), KEINE Anführungszeichen, KEINE
  Datumsbereiche, KEINE Klammern, KEIN Präfix wie "qmd query:".
- qmd kann nicht nach Datum filtern — gib also keine Datumsangaben mit.
- Beispiel-richtig: "Wolf Warnken Kündigung", "Coaching Eric Fischer".
- Beispiel-falsch: "'2024-06-20..2024-10-08' AND (X OR Y)".

Antworte als JSON: {{"reasoning": "...", "suggestions": ["...", ...]}}"""
    raw = chat(model, [{"role": "system", "content": SYSTEM_PROMPT},
                       {"role": "user", "content": prompt}], json_mode=True)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        print("Konnte Vorschlag nicht parsen, Rohausgabe:\n", raw)
        return
    print(f"\n{data.get('reasoning', '')}\n")
    sugg = data.get("suggestions", [])
    for i, q in enumerate(sugg, 1):
        print(f"  [{i}] {q}")
    print("\nMit  :1  :2  …  direkt ausführen, oder /search <text>.")
    state["last_suggestions"] = sugg


def cmd_reject(state: dict, pattern: str) -> None:
    """Bulk-Reject aller pending Mails, deren Betreff oder Teilnehmer
    den Pattern (case-insensitive Substring) enthalten."""
    if not pattern:
        print("Usage: /reject <pattern>   (matcht Betreff und Teilnehmer)")
        return
    pat = pattern.lower()
    matches = []
    for c in state["candidates"].values():
        if c["status"] != "pending":
            continue
        haystack = (c["subject"] + " " + " ".join(c["participants"])).lower()
        if pat in haystack:
            matches.append(c)
    if not matches:
        print(f"Keine pending Mail matcht {pattern!r}.")
        return
    print(f"\n{len(matches)} pending Mail(s) matchen {pattern!r}:")
    for c in sorted(matches, key=lambda x: x["date_start"]):
        parts = ", ".join(c["participants"][:2])
        print(f"  {c['date_start'] or '???':11s}  {c['subject'][:55]:<55s}  {parts}")
    try:
        ans = input(f"\nAlle {len(matches)} aussortieren? [j/N]: ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        print()
        return
    if ans not in {"j", "ja", "y", "yes"}:
        print("Abgebrochen.")
        return
    for c in matches:
        c["status"] = "rejected"
        c["rejection_reason"] = f"bulk-reject: {pattern}"
    print(f"✗ {len(matches)} Mail(s) auf 'rejected' gesetzt.")


def cmd_review(state: dict, model: str, out_dir: str,
               use_llm: bool = True) -> None:
    pending = [u for u, c in state["candidates"].items() if c["status"] == "pending"]
    if not pending:
        print("Keine Mails zum Review.")
        return
    print(f"\n{len(pending)} pending. (Strg-C bricht ab, Status bleibt erhalten.)")
    for uri in pending:
        c = state["candidates"][uri]
        print("\n" + "─" * 70)
        print(f"📧 {c['subject']}")
        date = c["date_start"] + (f" – {c['date_last']}"
                                  if c["date_last"] and c["date_last"] != c["date_start"]
                                  else "")
        print(f"   {date}  |  {c['message_count']} Nachricht(en)")
        print(f"   {', '.join(c['participants'][:3])}")
        print(f"   {uri}")

        llm = {"likely_relevant": None, "reasoning": "", "summary": ""}
        if use_llm:
            prompt = f"""Mail-Frontmatter:
- Betreff: {c['subject']}
- Datum: {c['date_start']}
- Teilnehmer: {', '.join(c['participants'])}

Auszug:
{c['body_excerpt']}

Fall: {state['case_description']}
{context_block(state, out_dir)}
Bewerte Relevanz und formuliere – falls relevant – eine juristisch
nüchterne Kernaussage (1 Satz, deutsch).
JSON: {{"likely_relevant": true|false, "reasoning": "...", "summary": "..."}}"""
            try:
                llm = json.loads(chat(model, [
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": prompt},
                ], json_mode=True))
            except KeyboardInterrupt:
                save_session(state)
                print("\nReview abgebrochen (während LLM-Anfrage). "
                      "Session gespeichert, Thread bleibt pending.")
                return
            except (json.JSONDecodeError, urllib.error.URLError) as e:
                print(f"   (LLM-Fehler: {e})")

            verdict = ("wahrscheinlich RELEVANT" if llm.get("likely_relevant")
                       else "wahrscheinlich nicht relevant")
            print(f"\n💡 {verdict}")
            if llm.get("reasoning"):
                print(f"   {llm['reasoning']}")
            if llm.get("summary"):
                print(f"   Kernaussage: {llm['summary']}")
        else:
            print(f"\n   {c['body_excerpt'][:600]}…")

        while True:
            try:
                ans = input("\n[a]kzeptieren / [r]aus / [s]kip / [b]ody / [q]uit: ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                save_session(state)
                print("\nReview abgebrochen, Session gespeichert.")
                return
            if ans == "b":
                print("\n" + qmd.fetch(uri, full=True)[:6000])
                continue
            if ans == "q":
                save_session(state)
                return
            if ans == "s":
                break
            if ans == "a":
                default = llm.get("summary", "")
                summary = input(f"Kernaussage [{default}]: ").strip() or default
                c["summary"] = summary
                c["status"] = "accepted"
                break
            if ans == "r":
                c["rejection_reason"] = input("Grund (optional): ").strip()
                c["status"] = "rejected"
                break
            print("Optionen: a r s b q")
        save_session(state)
    print("\nReview abgeschlossen.")


def cmd_list(state: dict) -> None:
    by = {"accepted": [], "pending": [], "rejected": []}
    for c in state["candidates"].values():
        by[c["status"]].append(c)
    for status, label in [("accepted", "✓ Akzeptiert"),
                          ("pending", "? Pending"),
                          ("rejected", "✗ Aussortiert")]:
        rows = sorted(by[status], key=lambda c: c["date_start"])
        print(f"\n{label} ({len(rows)}):")
        for c in rows:
            extra = f" — {c['summary']}" if status == "accepted" and c["summary"] else ""
            print(f"  {c['date_start'] or '???':11s}  {c['subject'][:60]}{extra}")


def cmd_gaps(state: dict, model: str, out_dir: str) -> None:
    accepted = sorted(
        [c for c in state["candidates"].values() if c["status"] == "accepted"],
        key=lambda c: c["date_start"],
    )
    if len(accepted) < 2:
        print("Zu wenige bestätigte Mails für Lückenanalyse.")
        return
    timeline = "\n".join(f"- {c['date_start']}: {c['subject']} — {c['summary']}"
                        for c in accepted)
    prompt = f"""Fall: {state['case_description']}
{context_block(state, out_dir)}
Bestätigte Mails:
{timeline}

Identifiziere zeitliche Lücken (>4 Wochen ohne Mails) oder thematische
Brüche, die auffällig sind. Schlage pro Lücke eine konkrete qmd-Suchanfrage
vor.

WICHTIG zum Feld "search":
- NUR reiner Stichworttext (2-5 deutsche Wörter / Eigennamen).
- KEINE Boolean-Operatoren (AND/OR), KEINE Anführungszeichen, KEINE
  Datumsbereiche, KEINE Klammern, KEIN Präfix wie "qmd query:".
- qmd kann nicht nach Datum filtern. Datum gehört ins "period"-Feld,
  NICHT in "search".
- Beispiel-richtig: "Lohnzahlung Verwaltungspflichten Geschäftsführer".
- Beispiel-falsch: "'2024-06-20..2024-10-08' AND (Lohn OR Abmahnung)".

JSON: {{"gaps": [{{"period": "...", "concern": "...", "search": "..."}}]}}"""
    try:
        data = json.loads(chat(model, [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ], json_mode=True))
    except json.JSONDecodeError as e:
        print(f"Parse-Fehler: {e}")
        return
    for g in data.get("gaps", []):
        print(f"\n• Zeitraum: {g.get('period', '?')}")
        print(f"  Bedenken: {g.get('concern', '')}")
        print(f"  Suche:    {g.get('search', '')}")


CLASS_MARK = {"confirmed": "✓", "extends": "+", "contradicts": "✗",
              "new": "—", "?": "?"}


def write_xlsx(path: Path, validations: list[tuple[dict, dict]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Zeitlicher Ablauf"
    headers = ["Datum", "Eingang (Postfach)", "Beteiligte", "Ereignis",
               "Bezug", "Kontext-Verweis", "Anmerkung", "Beleg"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for c, llm in validations:
        ev_date = llm.get("event_date") or c["date_start"]
        sent_date = c["date_start"] if c["date_start"] != ev_date else ""
        ws.append([
            ev_date,
            sent_date,
            "; ".join(c["participants"]),
            llm.get("event") or c["summary"],
            CLASS_MARK.get(llm.get("classification", "?"), "?"),
            llm.get("context_reference", ""),
            (c.get("resolution_note") or "").strip(),
            Path(c["uri"]).name,
        ])
    widths = [12, 12, 30, 60, 7, 28, 35, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[5].alignment = Alignment(wrap_text=True, vertical="top")
        row[6].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(horizontal="center", vertical="top")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(path)


def md_to_docx(md_path: Path, docx_path: Path) -> bool:
    """md -> docx via pandoc. Liefert True bei Erfolg, False sonst."""
    if not shutil.which("pandoc"):
        return False
    try:
        subprocess.run(
            ["pandoc", "-f", "gfm", "--standalone",
             "--toc", "--toc-depth=2",
             str(md_path), "-o", str(docx_path)],
            check=True, capture_output=True,
        )
        return True
    except subprocess.CalledProcessError as e:
        print(f"      pandoc-Fehler: {e.stderr.decode(errors='ignore')}")
        return False


INTERNAL_SECTION_RE = re.compile(
    r"^##\s*Interne\s*Kontakte[^\n]*\n(.*?)(?=^##\s|\Z)",
    re.MULTILINE | re.DOTALL | re.IGNORECASE,
)


def parse_internal_contacts(state: dict, out_dir: str) -> list[str]:
    """Extrahiert Filter-Patterns aus der `## Interne Kontakte`-Sektion
    der kontext.md. Liefert lowercase-Strings für Substring-Match."""
    text = load_context_text(state, out_dir)
    if not text:
        return []
    m = INTERNAL_SECTION_RE.search(text)
    if not m:
        return []
    patterns = []
    for line in m.group(1).splitlines():
        line = line.strip().lstrip("-*•").strip()
        if not line or line.startswith("#"):
            continue
        patterns.append(line.lower())
    return patterns


def is_internal(candidate: dict, patterns: list[str]) -> bool:
    if not patterns:
        return False
    haystack = (candidate.get("subject", "") + " " +
                " ".join(candidate.get("participants", []))).lower()
    return any(p in haystack for p in patterns)


def cmd_akte(state: dict, model: str, out_dir: str) -> None:
    """Kuratierter Gerichts-Export: validiert Mails gegen kontext.md,
    filtert interne Korrespondenz automatisch raus."""
    out = Path(out_dir).expanduser()
    accepted = sorted(
        [c for c in state["candidates"].values() if c["status"] == "accepted"],
        key=lambda c: c["date_start"],
    )
    if not accepted:
        print("Keine akzeptierten Mails — nichts zu validieren.")
        return

    ctx_text = load_context_text(state, out_dir).strip()
    if not ctx_text:
        print("Keine kontext.md vorhanden — /akte braucht den Kontext.\n"
              "Mit /context edit anlegen.")
        return

    patterns = parse_internal_contacts(state, out_dir)
    if not patterns:
        print("⚠  Keine '## Interne Kontakte'-Sektion in kontext.md gefunden.\n"
              "   Lege sie an (Liste mit Namen/Domains), sonst kann /akte\n"
              "   keine internen Mails herausfiltern.")
        try:
            ans = input("Trotzdem fortfahren (alle Mails als extern)? [j/N]: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            return
        if ans not in {"j", "ja", "y", "yes"}:
            return

    external = [c for c in accepted if not is_internal(c, patterns)]
    internal = [c for c in accepted if is_internal(c, patterns)]
    print(f"\n→ {len(accepted)} akzeptierte Mails: "
          f"{len(external)} extern, {len(internal)} intern (gefiltert).")
    if not external:
        print("Keine externen Mails — Akte wäre leer. Abbruch.")
        return

    # Aufräumen + Verzeichnisstruktur anlegen
    dossier_dir = out / "dossier"
    mails_dir = dossier_dir / "mails"
    intern_path = out / "dossier_intern.md"   # bewusst AUSSERHALB von dossier/

    out.mkdir(parents=True, exist_ok=True)
    if dossier_dir.exists():
        shutil.rmtree(dossier_dir)
        print("   Vorhandenes dossier/ entfernt.")
    if intern_path.exists():
        intern_path.unlink()
    # Auch die alten flachen akte_*-Dateien aus früheren Versionen aufräumen
    for fname in ("akte_zeitlicher_ablauf.md", "akte_zeitlicher_ablauf.csv",
                  "akte_zeitlicher_ablauf.xlsx",
                  "akte_zusammenfassung.md", "akte_zusammenfassung.docx",
                  "akte_intern.md"):
        f = out / fname
        if f.exists():
            f.unlink()
    legacy_mails = out / "akte_mails"
    if legacy_mails.exists():
        shutil.rmtree(legacy_mails)
    dossier_dir.mkdir(parents=True)
    mails_dir.mkdir(parents=True)

    # Validierung pro Mail
    print(f"\n→ Validiere {len(external)} externe Mails gegen kontext.md")
    print("   (Strg-C bricht ab, bisherige Ergebnisse werden geschrieben.)\n")
    validations: list[tuple[dict, dict]] = []
    for i, c in enumerate(external, 1):
        print(f"   [{i}/{len(external)}] {c['date_start']} — "
              f"{c['subject'][:55]}", flush=True)
        try:
            full_body = qmd.fetch(c["uri"], full=True)
        except subprocess.CalledProcessError as e:
            print(f"      qmd-Fehler: {e}")
            full_body = c["body_excerpt"]
        # Frontmatter abschneiden, nur Mail-Inhalt ans LLM geben
        _, body_only = qmd.parse_frontmatter(full_body)
        body_only = body_only[:6000]

        prompt = f"""Du erstellst einen Eintrag für die Gerichtsakte aus einer E-Mail
und validierst sie zugleich gegen die Erinnerung des Mandanten.

Erinnerung des Mandanten (kontext.md):
---
{ctx_text}
---

E-Mail:
- Versanddatum (Frontmatter): {c['date_start']}
- Betreff: {c['subject']}
- Beteiligte: {', '.join(c['participants'])}

Mail-Inhalt:
---
{body_only}
---

EREIGNISDATUM: Wenn der Betreff mit "WG:", "FW:", "Fwd:", "Wtr:"
beginnt oder das Mail-Layout mehrere Header zeigt ("Von: ...
Gesendet: ..."), dann ist das relevante Datum das der ursprünglichen
Mail im weitergeleiteten Block — NICHT das Versanddatum. Extrahiere
es aus dem Inhalt. Sonst Versanddatum.

EREIGNIS-FORMULIERUNG (Feld "event"):
- Ein einziger, vollständiger deutscher Satz für die Gerichtstabelle.
- Beschreibe konkret WAS geschehen ist, WER gehandelt hat (Person/Rolle),
  was der INHALTLICHE Kern war.
- Geschäftsmäßiger, juristisch-nüchterner Ton — präsentabel vor Gericht.
- KEINE Wiederholung des Betreffs.
- KEINE Phrasen wie "wurde mitgeteilt, dass…" wenn der Inhalt klar genug ist.
- Falls eine Verpflichtung, Frist, Forderung, Zusage oder Ablehnung
  ausgesprochen wurde: das ausdrücklich nennen.

KLASSIFIKATION (Feld "classification") — Beziehung zur Erinnerung:
- "confirmed":   bestätigt eine konkrete Aussage der Erinnerung
- "extends":     erweitert/präzisiert die Erinnerung
- "contradicts": widerspricht einer Aussage der Erinnerung
- "new":         Punkt, der in der Erinnerung gar nicht vorkommt

Wenn die Mail eine Aussage der Erinnerung direkt belegt/widerlegt,
nenne den Bezug knapp im Feld "context_reference" (sonst leer).

JSON: {{"classification": "confirmed|extends|contradicts|new",
       "event_date": "YYYY-MM-DD",
       "event": "<1 vollständiger Satz>",
       "context_reference": "<Bezug oder leer>"}}"""
        try:
            llm = json.loads(chat(model, [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ], json_mode=True))
        except KeyboardInterrupt:
            print("\n   Abgebrochen. Bisherige Ergebnisse werden geschrieben.")
            break
        except (json.JSONDecodeError, urllib.error.URLError) as e:
            print(f"      LLM-Fehler: {e}")
            llm = {"classification": "?", "event": c["summary"],
                   "context_reference": "", "event_date": ""}
        ed = (llm.get("event_date") or "").strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", ed):
            ed = c["date_start"]
        llm["event_date"] = ed
        validations.append((c, llm))

    validations.sort(key=lambda v: (v[1].get("event_date") or v[0]["date_start"]))

    if not validations:
        print("Keine validierten Mails. Abbruch.")
        return

    def esc(s: str) -> str:
        return str(s).replace("|", "\\|").replace("\n", " ")

    md = ["# Akte: Zeitlicher Ablauf", "",
          f"_Stand: {datetime.now().strftime('%Y-%m-%d')}, "
          f"{len(validations)} externe Mail(s) als Belege_", "",
          "**Bezug zur Erinnerung des Mandanten:** "
          "✓ bestätigt &nbsp; + erweitert &nbsp; ✗ widerspricht &nbsp; — neu",
          "",
          "_Datum = Datum des eigentlichen Ereignisses; "
          "Eingang = Datum der Mail im Postfach (bei Weiterleitungen abweichend)._",
          "",
          "| Datum | Eingang | Beteiligte | Ereignis | Bezug | Anmerkung | Beleg |",
          "|-------|---------|------------|----------|:----:|-----------|-------|"]
    for c, llm in validations:
        parts = "; ".join(c["participants"][:3])
        cls = llm.get("classification", "?")
        mark = CLASS_MARK.get(cls, "?")
        ref = (llm.get("context_reference") or "").strip()
        ref_suffix = f" _({esc(ref)})_" if ref else ""
        ev_date = llm.get("event_date") or c["date_start"]
        sent_date = c["date_start"] if c["date_start"] != ev_date else ""
        note = esc((c.get("resolution_note") or "").strip())
        md.append(
            f"| {ev_date} | {sent_date} | {esc(parts)} | "
            f"{esc(llm.get('event') or c['summary'])}{ref_suffix} | "
            f"{mark} | {note} | {Path(c['uri']).name} |"
        )
    (dossier_dir / "zeitlicher_ablauf.md").write_text("\n".join(md) + "\n")

    with (dossier_dir / "zeitlicher_ablauf.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Datum", "Eingang (Postfach)", "Beteiligte", "Ereignis",
                    "Bezug", "Kontext-Verweis", "Anmerkung", "Beleg"])
        for c, llm in validations:
            ev_date = llm.get("event_date") or c["date_start"]
            sent_date = c["date_start"] if c["date_start"] != ev_date else ""
            w.writerow([
                ev_date,
                sent_date,
                "; ".join(c["participants"]),
                llm.get("event") or c["summary"],
                llm.get("classification", "?"),
                llm.get("context_reference", ""),
                (c.get("resolution_note") or "").strip(),
                Path(c["uri"]).name,
            ])

    for c, _ in validations:
        body = qmd.fetch(c["uri"], full=True)
        (mails_dir / Path(c["uri"]).name).write_text(body)

    if internal:
        intern_md = [
            "# Aussortiert: interne Korrespondenz",
            "",
            f"_NUR FÜR DICH — bewusst außerhalb des dossier-Verzeichnisses, "
            f"damit es beim Zippen nicht versehentlich mitgeht. "
            f"{len(internal)} Mail(s), basierend auf `## Interne Kontakte` "
            f"in kontext.md._",
            "",
            f"Filter-Patterns: {', '.join(repr(p) for p in patterns)}",
            "",
            "| Datum | Beteiligte | Betreff | Datei |",
            "|-------|------------|---------|-------|",
        ]
        for c in sorted(internal, key=lambda x: x["date_start"]):
            parts = "; ".join(c["participants"][:3])
            intern_md.append(
                f"| {c['date_start']} | {esc(parts)} | "
                f"{esc(c['subject'])} | {Path(c['uri']).name} |"
            )
        intern_path.write_text("\n".join(intern_md) + "\n")

    print(f"\n→ Generiere narrative Zusammenfassung\n")
    def _line(c: dict, llm: dict) -> str:
        base = (f"- {llm.get('event_date') or c['date_start']} "
                f"[{CLASS_MARK.get(llm.get('classification','?'),'?')}] "
                f"{llm.get('event') or c['summary']}")
        note = (c.get("resolution_note") or "").strip()
        if note:
            base += f"\n  Klärung des Mandanten: {note}"
        return base

    timeline = "\n".join(_line(c, llm) for c, llm in validations)

    extra_resolutions = state.get("context_resolutions") or []
    extra_block = ""
    if extra_resolutions:
        extra_block = ("\n\nWeitere Klärungen des Mandanten (ohne direkten "
                       "Mail-Bezug):\n" +
                       "\n".join(f"- Behauptung: {r.get('claim', '')}\n"
                                 f"  Klarstellung: {r.get('note', '')}"
                                 for r in extra_resolutions))

    summary_prompt = f"""Mandant: {state['case_description']}

Erinnerung des Mandanten:
---
{ctx_text}
---

Validierte Timeline (chronologisch, externe Mails, mit Bezug zur Erinnerung):
{timeline}{extra_block}

Schreibe eine narrative Zusammenfassung für das Gericht (deutsch,
juristisch nüchtern). Strukturiere nach erkennbaren Phasen. Mache
deutlich, wo Mails die Erinnerung bestätigen (✓), erweitern/präzisieren
(+), widersprechen (✗) oder einen neuen Punkt einführen (—). Bei
Widersprüchen, zu denen der Mandant eine Klärung notiert hat, übernimm
diese Klärung sachlich in die Darstellung. KEINE Erfindungen — nur was
aus den validierten Mails und den Klärungen ersichtlich ist. Verweise
auf Daten und Beteiligte konkret.

PFLICHT-FORMATIERUNG (das Dokument wird nach Word konvertiert):
- Verwende für jede Phase eine Markdown-Überschrift Ebene 2:
  `## Phase 1: <Titel>`, `## Phase 2: <Titel>`, …
- Innerhalb der Phasen für Ereignisse Bullet-Listen mit `-`.
- Daten und zentrale Personen in **fett** hervorheben.
- KEIN reines `**Phase 1:**` ohne `##`-Heading — das wäre nur Fettdruck,
  keine Überschrift.
- Keine Tabellen, keine Code-Blöcke."""
    text = chat_stream(model, [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": summary_prompt},
    ])
    legend = (
        "## Legende: Bezug der Mails zur Erinnerung des Mandanten\n\n"
        "| Symbol | Bedeutung |\n"
        "|:---:|---|\n"
        "| ✓ | Mail bestätigt eine Aussage der Erinnerung |\n"
        "| + | Mail erweitert oder präzisiert die Erinnerung |\n"
        "| ✗ | Mail widerspricht einer Aussage der Erinnerung |\n"
        "| — | Mail führt einen Punkt ein, der in der Erinnerung fehlt |\n\n"
        "---\n\n"
    )
    unsupported = state.get("context_unsupported") or []
    unsupported_block = ""
    if unsupported:
        rows = []
        for u in unsupported:
            cl = u.get("claim", "")
            sr = (u.get("search") or "").strip()
            suffix = f" — _Such-Vorschlag: {sr}_" if sr else ""
            rows.append(f"- {cl}{suffix}")
        unsupported_block = (
            "\n\n## Aus der Erinnerung noch unbelegt\n\n"
            "Folgende Punkte aus der Erinnerung des Mandanten wurden bei "
            "der letzten `/validate-context`-Prüfung nicht durch Mails "
            "im Archiv belegt:\n\n"
            + "\n".join(rows) + "\n"
        )

    summary_md_path = dossier_dir / "zusammenfassung.md"
    summary_md_path.write_text(
        f"# Akte: Zusammenfassung\n\n"
        f"_Stand: {datetime.now().strftime('%Y-%m-%d')}_\n\n"
        f"{legend}"
        f"{text}\n"
        f"{unsupported_block}"
    )

    xlsx_path = dossier_dir / "zeitlicher_ablauf.xlsx"
    if HAVE_OPENPYXL:
        write_xlsx(xlsx_path, validations)
    else:
        print("⚠  openpyxl fehlt — keine .xlsx-Datei. "
              "Installation: pip3 install openpyxl")

    docx_path = dossier_dir / "zusammenfassung.docx"
    docx_ok = md_to_docx(summary_md_path, docx_path)
    if not docx_ok:
        print("⚠  pandoc nicht gefunden oder Fehler — keine .docx-Datei. "
              "Installation: brew install pandoc")

    print(f"\n📂 {dossier_dir}/  (alles für die Akte — bereit zum Zippen)")
    print(f"   ✓ zeitlicher_ablauf.md")
    print(f"   ✓ zeitlicher_ablauf.csv")
    if HAVE_OPENPYXL:
        print(f"   ✓ zeitlicher_ablauf.xlsx")
    print(f"   ✓ zusammenfassung.md")
    if docx_ok:
        print(f"   ✓ zusammenfassung.docx")
    print(f"   ✓ mails/  ({len(validations)} Belege)")
    if internal:
        print(f"\n⊘ {intern_path}")
        print(f"   (NUR für dich — {len(internal)} interne Mails aussortiert; "
              f"liegt bewusst AUSSERHALB von dossier/, damit es beim Zippen "
              f"nicht mit reingeht)")

    open_conflicts = [(c, llm) for c, llm in validations
                      if llm.get("classification") == "contradicts"
                      and not (c.get("resolution_note") or "").strip()]
    if open_conflicts:
        print(f"\n⚠  {len(open_conflicts)} OFFENE Widerspruch/Widersprüche "
              f"(noch nicht geklärt):")
        for c, _ in open_conflicts:
            print(f"   • {c['date_start']} — {c['subject'][:60]}")
        print("   Vor Übergabe klären mit /validate-context.")

    if unsupported:
        print(f"\n⚠  {len(unsupported)} Punkt(e) aus der Erinnerung sind "
              f"NOCH UNBELEGT:")
        for u in unsupported[:5]:
            print(f"   • {u.get('claim', '')[:70]}")
        if len(unsupported) > 5:
            print(f"   … und {len(unsupported) - 5} weitere "
                  f"(siehe zusammenfassung.md → 'Aus der Erinnerung noch unbelegt').")
        print("   Mit gezieltem /search oder /devil nachhaken.")
    elif state.get("context_unsupported") is None:
        print("\nℹ  Tipp: /validate-context prüft auch die Gegenrichtung — "
              "welche Punkte deiner Erinnerung von keiner Mail belegt sind. "
              "Lauf das vor /dossier, dann erscheinen sie in der Zusammenfassung.")


def cmd_export(state: dict, out_dir: str) -> None:
    out = Path(out_dir).expanduser()
    accepted = sorted(
        [c for c in state["candidates"].values() if c["status"] == "accepted"],
        key=lambda c: c["date_start"],
    )
    if not accepted:
        print("Keine akzeptierten Mails — nichts zu exportieren.")
        return
    (out / "mails").mkdir(parents=True, exist_ok=True)

    def esc(s: str) -> str:
        return str(s).replace("|", "\\|").replace("\n", " ")

    md = ["# Zeitlicher Ablauf", "",
          f"_Stand: {datetime.now().strftime('%Y-%m-%d')}, "
          f"{len(accepted)} Mails_", "",
          "| Datum | Betreff | Teilnehmer | Kernaussage | Quelle |",
          "|-------|---------|------------|-------------|--------|"]
    for c in accepted:
        parts = "; ".join(c["participants"][:3])
        md.append(f"| {c['date_start']} | {esc(c['subject'])} | "
                  f"{esc(parts)} | {esc(c['summary'])} | "
                  f"{Path(c['uri']).name} |")
    (out / "zeitlicher_ablauf.md").write_text("\n".join(md) + "\n")

    with (out / "zeitlicher_ablauf.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Datum", "Betreff", "Teilnehmer", "Kernaussage", "Quelle"])
        for c in accepted:
            w.writerow([c["date_start"], c["subject"],
                        "; ".join(c["participants"]), c["summary"],
                        Path(c["uri"]).name])

    for c in accepted:
        body = qmd.fetch(c["uri"], full=True)
        (out / "mails" / Path(c["uri"]).name).write_text(body)

    print(f"\n✓ {out / 'zeitlicher_ablauf.md'}")
    print(f"✓ {out / 'zeitlicher_ablauf.csv'}")
    print(f"✓ {out / 'mails'}/  ({len(accepted)} Dateien)")


def cmd_summary(state: dict, out_dir: str, model: str) -> None:
    accepted = sorted(
        [c for c in state["candidates"].values() if c["status"] == "accepted"],
        key=lambda c: c["date_start"],
    )
    if not accepted:
        print("Keine akzeptierten Mails.")
        return
    timeline = "\n".join(
        f"- {c['date_start']} — {c['subject']} "
        f"({', '.join(c['participants'][:2])}): {c['summary']}"
        for c in accepted
    )
    prompt = f"""Fall: {state['case_description']}
{context_block(state, out_dir)}
Zeitliche Abfolge:
{timeline}

Schreibe eine narrative Zusammenfassung (deutsch, juristisch nüchtern).
Strukturiere nach erkennbaren Phasen. Markiere offene Lücken explizit.
Keine Erfindungen — nur was aus den Mails ersichtlich ist."""
    print("→ Generiere Zusammenfassung...\n")
    text = chat_stream(model, [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": prompt},
    ])
    out_path = Path(out_dir).expanduser() / "zusammenfassung.md"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(f"# Zusammenfassung\n\n{text}\n")
    print(f"\n✓ {out_path}")


def cmd_freeform(state: dict, model: str, question: str, num_ctx: int,
                 out_dir: str) -> None:
    accepted = sorted(
        [c for c in state["candidates"].values() if c["status"] == "accepted"],
        key=lambda c: c["date_start"],
    )
    ctx = "\n".join(f"- {c['date_start']}: {c['subject']} — {c['summary']}"
                    for c in accepted) or "(nichts)"
    chat_stream(model, [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content":
            f"Fall: {state['case_description']}"
            f"{context_block(state, out_dir)}"
            f"\nBestätigte Mails:\n{ctx}\n\nFrage: {question}"},
    ], num_ctx=num_ctx)


def cmd_validate_context(state: dict, model: str, out_dir: str) -> None:
    """Prüft kontext.md gegen die akzeptierte Timeline."""
    ctx_text = load_context_text(state, out_dir).strip()
    if not ctx_text:
        print("Keine Kontext-Datei vorhanden — nichts zu validieren.\n"
              "Mit /context edit anlegen.")
        return
    accepted = sorted(
        [c for c in state["candidates"].values() if c["status"] == "accepted"],
        key=lambda c: c["date_start"],
    )
    if not accepted:
        print("Noch keine akzeptierten Mails — Validierung braucht Belegmaterial.")
        return
    timeline = "\n".join(
        f"- [{Path(c['uri']).name}] {c['date_start']} — {c['subject']} "
        f"({', '.join(c['participants'][:2])}): {c['summary']}"
        for c in accepted
    )
    prompt = f"""Fall: {state['case_description']}

Subjektive Erinnerung des Mandanten (vor Jahren) — kann fehlerhaft sein:
---
{ctx_text}
---

Belegte Mails aus dem Mailarchiv (chronologisch, mit Filename in []):
{timeline}

Prüfe Punkt für Punkt aus der Erinnerung gegen die Mails. Ordne jede
prüfbare Aussage in genau eine der drei Kategorien:

- "supported": Aussage wird von mindestens einer Mail gestützt
  (nenne im "evidence"-Feld Datum + Betreff der Mail)
- "contradicted": eine Mail widerspricht der Erinnerung
  (nenne im "issue"-Feld den Widerspruch + Datum + Betreff;
   nenne im "mail_ref"-Feld den Filename aus den []-Klammern oben,
   wenn die widersprechende Mail eindeutig identifizierbar ist —
   sonst "" lassen;
   im "search"-Feld eine 2-5-Wort-Stichwortsuche zur Vertiefung)
- "unsupported": Aussage hat keinen Mail-Beleg
  (im "search"-Feld eine 2-5-Wort-Stichwortsuche, mit der man Belege
   versuchen könnte; KEINE Boolean-Operatoren, keine Datumsbereiche)

Vage Hintergrundaussagen ("X war wichtig", "Y war kompliziert) ignoriere.
Konzentriere dich auf konkrete Behauptungen (Daten, Handlungen, Aussagen).

JSON: {{
  "supported":     [{{"claim": "...", "evidence": "..."}}],
  "contradicted":  [{{"claim": "...", "issue": "...", "mail_ref": "...", "search": "..."}}],
  "unsupported":   [{{"claim": "...", "search": "..."}}]
}}"""
    try:
        data = json.loads(chat(model, [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ], json_mode=True))
    except (json.JSONDecodeError, urllib.error.URLError) as e:
        print(f"LLM-Fehler: {e}")
        return

    suggestions: list[str] = []
    sup = data.get("supported", [])
    con = data.get("contradicted", [])
    uns = data.get("unsupported", [])

    print(f"\n✓ Gestützt durch Mails ({len(sup)}):")
    for s in sup:
        print(f"  • {s.get('claim', '')}")
        if s.get("evidence"):
            print(f"      → Beleg: {s['evidence']}")

    print(f"\n✗ Widerspruch zwischen Erinnerung und Mails ({len(con)}):")
    for s in con:
        print(f"  • {s.get('claim', '')}")
        if s.get("issue"):
            print(f"      → Konflikt: {s['issue']}")
        if s.get("mail_ref"):
            print(f"      → Mail:    {s['mail_ref']}")
        if s.get("search"):
            print(f"      → Suche:   {s['search']}")
            suggestions.append(s["search"])

    print(f"\n? Unbelegt — keine Mails dazu ({len(uns)}):")
    for s in uns:
        print(f"  • {s.get('claim', '')}")
        if s.get("search"):
            print(f"      → Suche: {s['search']}")
            suggestions.append(s["search"])

    if suggestions:
        state["last_suggestions"] = suggestions
        print(f"\n→ {len(suggestions)} Such-Vorschläge: mit :1 :2 … direkt ausführen.")

    # Unsupported-Liste persistieren - /dossier liest sie, um die
    # Gegenrichtung (Erinnerung → Mails) in der Zusammenfassung mitzuführen.
    state["context_unsupported"] = [
        {"claim": s.get("claim", ""), "search": s.get("search", "")}
        for s in uns
    ]
    save_session(state)

    # Interaktive Klärung der Widersprüche
    if not con:
        return
    try:
        ans = input(f"\n{len(con)} Widerspruch/Widersprüche jetzt klären? [j/N]: ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        print()
        return
    if ans not in {"j", "ja", "y", "yes"}:
        print("Übersprungen — Klärungen kannst du jederzeit per /validate-context nachholen.")
        return

    by_filename = {Path(c["uri"]).name: c["uri"]
                   for c in state["candidates"].values()}

    for i, w in enumerate(con, 1):
        print("\n" + "─" * 70)
        print(f"[{i}/{len(con)}] Behauptung: {w.get('claim', '')}")
        if w.get("issue"):
            print(f"   Konflikt: {w['issue']}")
        ref = (w.get("mail_ref") or "").strip()
        m = re.search(r"[\w.-]+\.md", ref)
        ref_clean = m.group(0) if m else ""
        if ref_clean and ref_clean in by_filename:
            print(f"   Mail:     {ref_clean}  (zugeordnet)")
        elif ref:
            print(f"   Mail-Hinweis: {ref}  (nicht eindeutig zugeordnet)")
        else:
            print("   (keine Mail-Referenz vom LLM angegeben)")

        while True:
            try:
                act = input("\n  [k]ontext editieren / [e]rklären / "
                            "[s]kip / [q]uit: ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                save_session(state)
                print("\n  Klärung abgebrochen, Session gespeichert.")
                return
            if act == "q":
                save_session(state)
                return
            if act == "s":
                break
            if act == "k":
                cmd_context(state, "edit", out_dir)
                continue
            if act == "e":
                try:
                    note = input("  Anmerkung (1-2 Sätze): ").strip()
                except (EOFError, KeyboardInterrupt):
                    print()
                    break
                if not note:
                    print("  Leere Eingabe, übersprungen.")
                    break
                if ref_clean and ref_clean in by_filename:
                    uri = by_filename[ref_clean]
                    state["candidates"][uri]["resolution_note"] = note
                    print(f"  ✓ Anmerkung an Mail {ref_clean} gespeichert.")
                else:
                    state.setdefault("context_resolutions", []).append({
                        "claim": w.get("claim", ""),
                        "issue": w.get("issue", ""),
                        "note": note,
                    })
                    print("  ✓ Anmerkung als allgemeine Klärung gespeichert "
                          "(keine Mail-Zuordnung).")
                save_session(state)
                break
            print("  Optionen: k e s q")


def cmd_devil(state: dict, model: str, out_dir: str) -> None:
    """Anwalt der Gegenseite: Schwachstellen identifizieren + Such-Stichworte."""
    accepted = sorted(
        [c for c in state["candidates"].values() if c["status"] == "accepted"],
        key=lambda c: c["date_start"],
    )
    if not accepted:
        print("Keine akzeptierten Mails — Devil's Advocate braucht Material.")
        return
    timeline = "\n".join(
        f"- {c['date_start']} — {c['subject']} "
        f"({', '.join(c['participants'][:2])}): {c['summary']}"
        for c in accepted
    )
    prompt = f"""Fall: {state['case_description']}
{context_block(state, out_dir)}
Bisheriger Sachverhalt aus Sicht des Mandanten (chronologisch):
{timeline}

Du bist Anwalt der Gegenseite. Identifiziere die Schwachstellen, an
denen die Gegenpartei Joschas Position angreifen wird:
- Inkonsistenzen oder zeitliche Lücken
- Aussagen ohne Mail-Beleg
- Mehrdeutige Formulierungen
- Mögliche alternative Lesarten der Vorgänge

Pro Schwachstelle: kurze Begründung + KONKRETE qmd-Stichwortsuche
(2-5 deutsche Wörter, KEINE Boolean-Operatoren, KEINE Quotes, KEINE
Datumsbereiche), mit der Joscha gezielt nach Belegen suchen kann.

JSON: {{"summary": "...", "weaknesses": [
  {{"point": "...", "explanation": "...", "search": "..."}}, ...
]}}"""
    try:
        data = json.loads(chat(model, [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ], json_mode=True))
    except (json.JSONDecodeError, urllib.error.URLError) as e:
        print(f"LLM-Fehler: {e}")
        return
    print(f"\n😈  Devil's Advocate:\n{data.get('summary', '')}\n")
    suggestions: list[str] = []
    for i, w in enumerate(data.get("weaknesses", []), 1):
        print(f"[{i}] {w.get('point', '')}")
        if w.get("explanation"):
            print(f"    {w['explanation']}")
        s = w.get("search", "").strip()
        if s:
            print(f"    → Suche: {s}")
            suggestions.append(s)
        print()
    if suggestions:
        state["last_suggestions"] = suggestions
        print("Mit  :1  :2  …  direkt ausführen.")


SEARCH_INTENT_RE = re.compile(
    r"^\s*(?:suche?|finde?|search|find)(?:\s+(?:nach|mal))?\s+(.+?)\s*[.!?]?$",
    re.IGNORECASE,
)


HELP = """
Befehle:
  /search <query>           qmd-Suche, fügt Treffer als 'pending' hinzu
  /add <uri> [<uri>...]     URIs direkt aufnehmen (oder /add @datei.txt)
  /suggest                  LLM schlägt 3-5 Suchanfragen vor
  :1  :2  …                 letzten Vorschlag Nr. N direkt ausführen
  /review                   durch pending-Mails iterieren (a/r/s/b/q)
  /review fast              dito, aber ohne LLM-Hypothese (manuell, schnell)
  /reject <pattern>         alle pending mit Pattern in Betreff/Teilnehmer
                            bulk-aussortieren (mit Bestätigung)
  /list                     Status-Übersicht der aktuellen Session
  /sessions                 alle vorhandenen Sessions auflisten
  /gaps                     LLM-Lückenanalyse zwischen bestätigten Mails
  /devil                    Anwalt der Gegenseite — Schwachstellen + Suchen
  /context [<pfad>|edit]    Hintergrund-Markdown anzeigen / setzen / öffnen
  /validate-context         prüft kontext.md gegen akzept. Mails (Konflikte!)
  /summary                  narrative Zusammenfassung erzeugen
  /export                   roher Dump (alle accepted) → Markdown+CSV+mails/
  /dossier                  kuratierter Gerichts-Export: validiert Mails
                            gegen kontext.md, filtert interne Korrespondenz
                            (laut '## Interne Kontakte' in kontext.md) raus
                            (Alias: /akte)
  /case [<text>]            Fallbeschreibung anzeigen / setzen
  /edit <subject-fragment>  Kernaussage einer akzept. Mail ändern
  /undo <subject-fragment>  Mail zurück auf 'pending'
  /help                     diese Hilfe
  /quit                     Beenden (Session ist auto-gespeichert)

"Suche nach X" / "Finde X"  →  qmd-Suche (wie /search X).
Sonstiger freier Text       →  Frage an das LLM mit aktuellem Stand als Kontext.
"""


def main() -> None:
    p = argparse.ArgumentParser(
        description="Interaktiver Mail-Recherche-Agent (qmd + ollama).")
    p.add_argument("-s", "--session",
                   help="Session-Name. Pflicht außer bei --list-sessions.")
    p.add_argument("--list-sessions", action="store_true",
                   help="Vorhandene Sessions auflisten und beenden.")
    p.add_argument("-m", "--model", default=DEFAULT_MODEL)
    p.add_argument("-n", "--top-k", type=int, default=20,
                   help="qmd Top-N pro Suche (Default 20).")
    p.add_argument("-o", "--out", default=str(DEFAULT_OUT),
                   help=f"Output-Verzeichnis (Default {DEFAULT_OUT}).")
    p.add_argument("-c", "--num-ctx", type=int,
                   default=int(os.environ.get("OLLAMA_NUM_CTX", "32768")))
    args = p.parse_args()

    if args.list_sessions:
        print_sessions()
        return
    if not args.session:
        p.error("--session/-s ist Pflicht (außer mit --list-sessions).")

    setup_readline(SESSIONS_DIR / ".recherche_history")

    state = load_session(args.session)
    if not state["case_description"]:
        print("Kurze Fallbeschreibung (1-3 Sätze, später per /case änderbar):")
        try:
            state["case_description"] = input("> ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            return
        save_session(state)

    counts = status_counts(state)
    print(f"\nMail-Recherche — Modell: {args.model}, Session: {args.session}")
    print(f"Output: {args.out}")
    print(f"Status: {counts['accepted']} akzept., {counts['pending']} pending, "
          f"{counts['rejected']} aussortiert. /help für Befehle.\n")

    while True:
        try:
            line = input("⚖  ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break
        if not line:
            continue
        try:
            if line in {"/quit", "/exit"}:
                break
            if line == "/help":
                print(HELP); continue
            if line == "/list":
                cmd_list(state); continue
            if line == "/sessions":
                print_sessions()
                print(f"(Aktuell: {args.session}. Wechseln = /quit + neu starten "
                      f"mit -s ANDERER_NAME.)")
                continue
            if line.startswith("/case"):
                rest = line[len("/case"):].strip()
                if rest:
                    state["case_description"] = rest
                    save_session(state)
                    print("Fallbeschreibung aktualisiert.")
                else:
                    print(state["case_description"] or "(leer)")
                continue
            if line.startswith("/search"):
                cmd_search(state, line[len("/search"):].strip(), args.top_k)
                save_session(state); continue
            if line.startswith("/add"):
                cmd_add(state, line[len("/add"):].strip())
                save_session(state); continue
            if line == "/suggest":
                cmd_suggest(state, args.model, args.out)
                save_session(state); continue
            if line == "/devil":
                cmd_devil(state, args.model, args.out)
                save_session(state); continue
            if line in {"/validate-context", "/validate"}:
                cmd_validate_context(state, args.model, args.out)
                save_session(state); continue
            if line.startswith("/context"):
                cmd_context(state, line[len("/context"):].strip(), args.out)
                save_session(state); continue
            if line.startswith(":") and line[1:].isdigit():
                idx = int(line[1:]) - 1
                sugg = state.get("last_suggestions", [])
                if 0 <= idx < len(sugg):
                    cmd_search(state, sugg[idx], args.top_k)
                    save_session(state)
                else:
                    print("Ungültige Nummer.")
                continue
            if line == "/review":
                cmd_review(state, args.model, args.out); continue
            if line in {"/review fast", "/review-fast"}:
                cmd_review(state, args.model, args.out, use_llm=False); continue
            if line.startswith("/reject"):
                cmd_reject(state, line[len("/reject"):].strip())
                save_session(state); continue
            if line == "/gaps":
                cmd_gaps(state, args.model, args.out); continue
            if line == "/summary":
                cmd_summary(state, args.out, args.model); continue
            if line == "/export":
                cmd_export(state, args.out); continue
            if line in {"/dossier", "/akte"}:
                cmd_akte(state, args.model, args.out); continue
            if line.startswith("/edit"):
                frag = line[len("/edit"):].strip().lower()
                hits = [c for c in state["candidates"].values()
                        if c["status"] == "accepted" and frag in c["subject"].lower()]
                if not hits:
                    print("Kein Treffer.")
                    continue
                if len(hits) > 1:
                    print(f"Mehrdeutig ({len(hits)} Treffer), bitte präziser.")
                    continue
                c = hits[0]
                new = input(f"Neue Kernaussage [{c['summary']}]: ").strip()
                if new:
                    c["summary"] = new
                    save_session(state)
                continue
            if line.startswith("/undo"):
                frag = line[len("/undo"):].strip().lower()
                for c in state["candidates"].values():
                    if c["status"] != "pending" and frag in c["subject"].lower():
                        c["status"] = "pending"
                        c["summary"] = ""
                        c["rejection_reason"] = ""
                        save_session(state)
                        print(f"→ {c['subject']} zurück auf pending.")
                        break
                else:
                    print("Kein Treffer.")
                continue
            m = SEARCH_INTENT_RE.match(line)
            if m:
                cmd_search(state, m.group(1).strip(), args.top_k)
                save_session(state)
                continue
            cmd_freeform(state, args.model, line, args.num_ctx, args.out)
        except urllib.error.URLError as e:
            print(f"Verbindung zu ollama fehlgeschlagen: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()
